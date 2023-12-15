import subprocess
import time
import pandas as pd
import shutil
import os
import matplotlib.pyplot as plt
import numpy as np 
from scipy.interpolate import make_interp_spline
import csv
import openpyxl
import warnings

warnings.filterwarnings("ignore")


# Create a DataFrame with the specified columns and data


def combine_sheets(file_list, output_file='combined_excel.xlsx'):
    # Dictionary to keep track of dataframes for each sheet name
    sheets_dict = {}

    for file_index, file_name in enumerate(file_list):
        xls = pd.ExcelFile(file_name)
        for sheet_name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name)
            df['Source'] = f'File_{file_index + 1}'  # Add a column to indicate the source file
            df = pd.concat([df.iloc[:, :1], df.iloc[:, 1:].reindex(sorted(df.columns[1:]), axis=1)], axis=1)
            # Append data to the respective sheet's dataframe in the dictionary
            if sheet_name not in sheets_dict:
                sheets_dict[sheet_name] = df
            else:
                sheets_dict[sheet_name] = pd.concat([sheets_dict[sheet_name], df], ignore_index=True)

    # Write each dataframe to its corresponding sheet in the combined Excel file
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for sheet_name, dataframe in sheets_dict.items():
            dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
        

def calculate_averages(file_path, group_name_prefix='(Group: ', group_number=8, total_rounds=10):
    # Open the Excel file
    xls = pd.ExcelFile(file_path)

    # Initialize lists to store the average values for lifetime, points, and energy for each round for the specified group
    group_lifetime_averages = []
    group_points_averages = []
    group_energy_averages = []

    # Generate the column names for the group
    # Assuming 12 agents per group
    group_name = f'{group_name_prefix}{group_number}'
    # Process each sheet and calculate averages
    for sheet_name in ['Lifetime', 'Points Average', 'Energy Average']:
        if sheet_name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name, index_col='Round')
            group_cols = [col for col in df.columns if group_name in str(col)]
            averages_list = []

            # Calculate the average for the specified group across the specified rounds
            for round_num in range(1, total_rounds + 1):
                if round_num in df.index:
                    averages_list.append(df.loc[round_num, group_cols].mean())

            # Assign the calculated averages to the respective lists
            if sheet_name == 'Lifetime':
                group_lifetime_averages = averages_list
            elif sheet_name == 'Points Average':
                group_points_averages = averages_list
            elif sheet_name == 'Energy Average':
                group_energy_averages = averages_list

    return group_lifetime_averages, group_points_averages, group_energy_averages

# Path to the directory containing the GoLang script and statistics.xlsx
# golang_script_directory = r"C:\Users\rohan\OneDrive - Imperial College London\Imperial Year 4\SOMAS\experiments\SOMAS2023"
golang_script_directory = r"/Users/chenchuwen/SOMAS/SOMAS2023Group8"
data_directory = r"/Users/chenchuwen/SOMAS/SOMAS2023Group8/experiment_data"
all_data = []
games = 100  # Specify the number of iterations
original_file_name = 'statistics.xlsx'
completed = 0

while completed < games:
    print(f"==================== Iteration {completed + 1} ====================")
    excel_files = []
    # Run the GoLang script that generates statistics.xlsx
    try:
        result = subprocess.run(['go', 'run', 'main.go'], cwd=golang_script_directory)
        if result.returncode == 0: 
            
        ##------------------------------ names need to be changed ---------------------------------------------
            title = "Final Performance"
            xlsx_file = f"experiment_data/Final_Performance_{completed+1}.xlsx"
        ##-----------------------------------------------------------------------------------------------------
        
            # Copy and rename the newly generated statistics.xlsx to a new name
            new_file_name = f'statistics{completed+1}.xlsx'
            new_file_path = os.path.join(golang_script_directory, new_file_name)
            shutil.copyfile(os.path.join(golang_script_directory, original_file_name), new_file_path)
            excel_files.append(new_file_path)

            time.sleep(1)  # Wait a bit before the next iteration (if necessary)
            completed += 1
            lifetime_group8,points_group8,energy_group8= calculate_averages(new_file_path)
            lifetime_group0,points_group0,energy_group0= calculate_averages(new_file_path,group_number=0)
            lifetime_group1,points_group1,energy_group1= calculate_averages(new_file_path,group_number=1)
            lifetime_group2,points_group2,energy_group2= calculate_averages(new_file_path,group_number=2)
            lifetime_group4,points_group4,energy_group4= calculate_averages(new_file_path,group_number=4)
            lifetime_group5,points_group5,energy_group5= calculate_averages(new_file_path,group_number=5)
            lifetime_group7,points_group7,energy_group7= calculate_averages(new_file_path,group_number=7)
            lifetime_other_groups=np.mean([lifetime_group1,lifetime_group2,lifetime_group4,lifetime_group5,lifetime_group7],axis=0)
            points_other_groups=np.mean([points_group1,points_group2,points_group4,points_group5,points_group7],axis=0)
            energy_other_groups=np.mean([energy_group1,energy_group2,energy_group4,energy_group5,energy_group7],axis=0)
            average_liftimes=np.mean(lifetime_group8)
            average_points=np.mean(points_group8)
            average_energy=np.mean(energy_group8)
            average_lifetimes_base=np.mean(lifetime_group0)
            averages_points_base=np.mean(points_group0)
            average_energy_base=np.mean(energy_group0)
            columns = ['Average_Lifetimes', 'Average_Points', 'Average_energy', 'Averages_Lifetimes_Base', 'Averages_Points_Base', 'Average_energy_base',
                       'Averages_Lifetimes_Other', 'Averages_Points_Other', 'Average_energy_Other']
            all_data.append(columns)
            for i in range(len(lifetime_group8)):
                all_data.append([lifetime_group8[i], points_group8[i], energy_group8[i], lifetime_group0[i], points_group0[i], energy_group0[i],
                                 lifetime_other_groups[i], points_other_groups[i], energy_other_groups[i]])
            ## all_data.append([average_liftimes, average_points, average_energy, average_lifetimes_base, averages_points_base, average_energy_base])
            print(f"Average lifetime for agent 8 is {average_liftimes}")
            print(f"Average points for agent 8 is {average_points}")
            print(f"Average lifetime for agent 0 is {average_lifetimes_base}")
            print(f"Average points for agent 0 is {averages_points_base}")
            
            
            
            ###plotting
            # First, we must interpolate the data points for each group to create smooth curves for plotting

        

            # Assuming the following data is available from your calculations
            # For demonstration, I will create some random data
            rounds = np.arange(1, 11)  # Rounds 1 to 10
            

            # Define colors for the plot lines
            grid_color = (0.9, 0.9, 0.9)
            lighter_gray_color = (0.8, 0.8, 0.8)
            colors = ['salmon','skyblue',lighter_gray_color]

            # Lifetime plot
            fig, axs = plt.subplots(1, 2, figsize=(18, 5))
            
            for ax in axs:
                ax.grid(True, color=grid_color, linestyle='-', linewidth=0.5)
                
            
            fig.suptitle(title, fontsize=16)
            axs[0].plot(rounds, lifetime_group8, label='Agent 8', marker='o', linestyle='-', color=colors[0], lw=3,markersize=8)
            axs[0].plot(rounds, lifetime_group0, label='BaseBiker',marker='o', linestyle='-', color=colors[1], lw=3,markersize=8)
            axs[0].plot(rounds, lifetime_other_groups,label='All Other Groups',linestyle='-', color=colors[2], lw=3,markersize=8)
            axs[0].set_title('Lifetime Comparison')
            axs[0].set_xlabel('Round')
            axs[0].set_ylabel('Lifetime')
            axs[0].set_ylim(0, 120)
            axs[0].set_xlim(1, 10)
            axs[0].set_yticks(np.arange(0, 121, 10))
            axs[0].set_xticks(np.arange(1, 11))
            axs[0].legend(loc='upper right',fontsize=7.5)
            axs[0].grid(True)

            # Points plot
            axs[1].plot(rounds, points_group8, label='Agent 8', marker='o', linestyle='-', color=colors[0], lw=3,markersize=8)
            axs[1].plot(rounds, points_group0, label='BaseBiker',marker='o', linestyle='-', color=colors[1], lw=3,markersize=8)
            axs[1].plot(rounds, points_other_groups,label='All Other Groups',linestyle='-', color=colors[2], lw=3,markersize=8)
            axs[1].set_title('Points Comparison')
            axs[1].set_xlabel('Round')
            axs[1].set_ylabel('Points')
            axs[1].set_xlim(1, 10)
            axs[1].set_ylim(0, 12)
            axs[1].set_yticks(np.arange(0, 13, 1))
            axs[1].set_xticks(np.arange(1, 11))
            axs[1].legend(loc='upper right',fontsize=7.5)
            axs[1].grid(True)


            # Energy plot
            # axs[2].plot(rounds, energy_group8, label='Agent 8', marker='o', linestyle='-', color=colors[0], lw=3,markersize=8)
            # axs[2].plot(rounds, energy_group0, label='BaseBiker', marker='o',linestyle='-', color=colors[1], lw=3,markersize=8)
            # axs[2].plot(rounds, energy_other_groups,label='All Other Groups',linestyle='-', color=colors[2], lw=3,markersize=8)
            # axs[2].set_title('Energy Comparison')
            # axs[2].set_xlabel('Round')
            # axs[2].set_ylabel('Energy')
            # axs[2].set_xlim(1, 10)
            # axs[2].set_ylim(0, 1.2)
            # axs[2].set_yticks(np.arange(0, 1.3, 0.1))
            # axs[2].set_xticks(np.arange(1, 11))
            # axs[2].legend(loc='upper right',fontsize=7.5)
            # axs[2].grid(True)
            plt.tight_layout()
            fig.savefig(os.path.join('experiment_image', f'{title}_{completed}.jpg'), dpi=800)  # Specify the folder path and desired file name
            new_data_df = pd.DataFrame(all_data, columns=columns)
            # 创建一个新的工作簿对象或加载已有的工作簿
            try:
                workbook = openpyxl.load_workbook(xlsx_file)
            except FileNotFoundError:
                workbook = openpyxl.Workbook()

            # 选择要操作的工作表，这里选择第一个工作表
            sheet = workbook.active

            # 将数据写入工作表，可以选择覆盖或追加
            # 如果要覆盖已有内容，首先清空工作表的内容
            if sheet.max_row > 1:
                sheet.delete_rows(2, sheet.max_row)
                
            # 然后将新数据写入工作表
            for row_data in all_data:
                sheet.append(row_data)

            # 保存工作簿
            workbook.save(xlsx_file)

            # if os.path.exists(csv_file) and os.stat(csv_file).st_size > 0:
            #     # Read existing data
            #     existing_data = pd.read_csv(csv_file)
            #     # Append ONLY the new data
            #     all_data_df = pd.concat([existing_data, new_data_df], ignore_index=True)
            # else:
            #     # If the file doesn't exist or is empty, use new data as all data
            #     all_data_df = new_data_df

            # # Save the updated data to the CSV file
            # # Write only the new data if the file exists and is not empty
            # if os.path.exists(csv_file) and os.stat(csv_file).st_size > 0:
            #     # Write only the new rows (difference between all_data_df and existing_data)
            #     new_rows_to_write = all_data_df.drop(existing_data.index)
            #     new_rows_to_write.to_csv(csv_file, mode='a', header=False, index=False)
            # else:
            #     # Write with header if the file is new or empty
            #     all_data_df.to_csv(csv_file, index=False)
            # plt.show()
            # csv_file = 'decide_force_energy_thres.csv'
            # columns = ['Average_Lifetimes', 'Average_Points', 'Averages_Lifetimes_Base', 'Averages_Points_Base']
            # new_data_df = pd.DataFrame(all_data, columns=columns)
            # if os.path.exists(csv_file) and os.stat(csv_file).st_size > 0:
            #     # Read existing data
            #     existing_data = pd.read_csv(csv_file)
            #     # Append ONLY the new data
            #     all_data_df = pd.concat([existing_data, new_data_df], ignore_index=True)
            # else:
            #     # If the file doesn't exist or is empty, use new data as all data
            #     all_data_df = new_data_df

            # # Save the updated data to the CSV file
            # # Write only the new data if the file exists and is not empty
            # if os.path.exists(csv_file) and os.stat(csv_file).st_size > 0:
            #     # Write only the new rows (difference between all_data_df and existing_data)
            #     new_rows_to_write = all_data_df.drop(existing_data.index)
            #     new_rows_to_write.to_csv(csv_file, mode='a', header=False, index=False)
            # else:
            #     # Write with header if the file is new or empty
            #     all_data_df.to_csv(csv_file, index=False)
        else:
            print("Error running the simulation. Rerunning the script.")
    except Exception as e:
        print("Error running the simulation. Rerunning the script.---------------------------------------")
        print(e)

# Combine the generated Excel files
combine_sheets(excel_files)
print("Excel sheets combined successfully.")  
